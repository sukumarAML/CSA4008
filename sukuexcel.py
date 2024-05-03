from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("firelog.xlsx")

ws = wb.active

#print(ws['A1'].value, ws['B2'].value)
      
for row in range(1,11):
    for col in range(1, 63):
        char = get_column_letter(col)
        print(ws[char+str(row)].value)

